import re
import aiohttp
import asyncio
import json
import subprocess
import sys
import importlib
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from rich.console import Console
from rich.progress import Progress, BarColumn, TimeElapsedColumn, TextColumn
from aiofiles import open as aio_open

required_packages = [
    'aiohttp', 
    'openpyxl', 
    'rich', 
    'aiofiles'
]

def install_requirements():
    console.print("[yellow]Проверка зависимостей...[/yellow]")
    for package in required_packages:
        try:
            importlib.import_module(package)
            console.print(f"[green]{package} уже установлен.[/green]")
        except ImportError:
            console.print(f"[red]{package} не найден. Установка...[/red]")
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])
            console.print(f"[green]{package} успешно установлен.[/green]")

with open("report_settings.json", "r") as settings_file:
    settings = json.load(settings_file)

output_file = settings["output_file"]
api_token = settings["api_token"]
telegram_token = settings["telegram_token"]
chat_id = settings["chat_id"]
max_concurrent_requests = settings["max_concurrent_requests"]

def get_api_url(ip: str) -> str:
    return f"https://api.findip.net/{ip}/?token={api_token}"

console = Console()

ban_regex = re.compile(r"(\d+\.\d+\.\d+\.\d+)")

banned_ips = defaultdict(lambda: {"country": None, "city": None, "provider": None})

async def fetch_ips_from_iptables():
    try:
        result = subprocess.run(
            ["iptables", "-L", "f2b-sshd", "-n", "--line-numbers"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )
        ip_lines = result.stdout.splitlines()
        for line in ip_lines:
            match = ban_regex.search(line)
            if match:
                ip = match.group(1)
                banned_ips[ip]
    except Exception as e:
        console.print(f"[red]Ошибка при получении данных из iptables: {e}[/red]")

async def fetch_ip_data(session, ip, progress_task, progress_bar):
    try:
        async with session.get(get_api_url(ip)) as response:
            if response.status == 200:
                data = await response.json()
                banned_ips[ip]["country"] = data["country"]["names"].get("en", "")
                banned_ips[ip]["city"] = data["city"]["names"].get("en", "")
                banned_ips[ip]["provider"] = data["traits"].get("isp", "")
    except Exception as e:
        console.print(f"[red]Ошибка для {ip}: {e}[/red]")
    finally:
        progress_bar.update(progress_task, advance=1)

async def process_ips():
    async with aiohttp.ClientSession(connector=aiohttp.TCPConnector(limit_per_host=max_concurrent_requests)) as session:
        with Progress(
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            "[progress.completed]{task.completed}/{task.total}",
            TimeElapsedColumn(),
            console=console,
        ) as bar:
            task = bar.add_task("Обработано", total=len(banned_ips))
            await asyncio.gather(*[fetch_ip_data(session, ip, task, bar) for ip in banned_ips])

def save_to_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "IP Report"

    headers = ["IP Address", "Country", "City", "Provider"]
    ws.append(headers)
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    sorted_ips = sorted(
        banned_ips.items(),
        key=lambda x: x[1].get("time", datetime.min),
    )

    for ip, data in sorted_ips:
        ws.append([
            ip,
            data.get("country", ""),
            data.get("city", ""),
            data.get("provider", ""),
        ])

    for col_num, col_cells in enumerate(ws.columns, start=1):
        col_width = max(len(str(cell.value or "")) for cell in col_cells) + 6
        ws.column_dimensions[get_column_letter(col_num)].width = col_width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(banned_ips) + 1}"
    wb.save(output_file)

async def send_stats_to_telegram(stats):
    stats_message = f"""
    🕒 *Статистика выполнения скрипта* 🕒
    
📅 *Дата и время начала*: {stats['start_time']}
⏳ *Время выполнения*: {stats['duration_seconds']} с
🌍 *Обработано IP*: {stats['total_ips']}

✨ Завершено успешно! ✨
    """
    async with aiohttp.ClientSession() as session:
        payload = {"chat_id": chat_id, "text": stats_message, "parse_mode": "Markdown"}
        async with session.post(f"https://api.telegram.org/bot{telegram_token}/sendMessage", json=payload) as response:
            if response.status != 200:
                console.print(f"[red]Ошибка отправки статистики: {await response.text()}[/red]")


async def send_file_to_telegram(file_path):
    async with aiohttp.ClientSession() as session:
        with open(file_path, "rb") as file:
            form_data = aiohttp.FormData()
            form_data.add_field("chat_id", chat_id)
            form_data.add_field("document", file, filename=file_path)
            async with session.post(f"https://api.telegram.org/bot{telegram_token}/sendDocument", data=form_data) as response:
                if response.status != 200:
                    console.print(f"[red]Ошибка отправки файла: {await response.text()}[/red]")


async def main():
    install_requirements()

    start_time = datetime.now()
    await fetch_ips_from_iptables()
    await process_ips()
    save_to_excel()

    duration = (datetime.now() - start_time).seconds
    stats = {
        "start_time": start_time.strftime("%Y-%m-%d %H:%M:%S"),
        "total_ips": len(banned_ips),
        "duration_seconds": duration,
    }
    await send_stats_to_telegram(stats)
    await send_file_to_telegram(output_file)


if __name__ == "__main__":
    asyncio.run(main())
